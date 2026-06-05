"""
=============================================================
  Аналіз бронь — ЗП  |  Все в одному файлі  (v5)
=============================================================
Запуск:  python bron_analysis_all.py
=============================================================

Copyright (C) 2026 r0665629590-droid

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, version 3.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
GNU General Public License for more details: https://www.gnu.org/licenses/gpl-3.0.html

Для комерційного використання без обмежень GPL —
звертайтесь до автора через GitHub.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os, glob, json, csv, re
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ──────────────────── НАЛАШТУВАННЯ ─────────────────────────
OUTPUT_SHEET    = "Аналіз бронь"
DEFAULT_MIN_SAL = 8_647
SETTINGS_FILE   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bron_settings.json")

DEFAULT_COEFFS = [
    {"name": "Звичайні (×2.5)",  "value": 2.5},
    {"name": "Медики (×4.0)",    "value": 4.0},
]

CANDIDATE_COLS = {
    "employee": ["Співробітник", "Фізична особа"],
    "dept":     ["Підрозділ"],
    "amount":   ["Сума взаєморозрахунків", "Сума упр", "Сума"],
    "group":    ["Група нарахування утримання виплати", "Код операції"],
}
FILTER_CANDIDATES = ["Нараховано", "Нарахування"]
BRON_YES_VALUES   = {"так", "yes", "true", "+", "1"}
STATUS_OK         = "OK — поріг виконано"
STATUS_NEED       = "ПОТРІБНО ДОНАРАХУВАТИ"

# Статуси з файлу Дії
DIA_EXCLUDED    = "Виключено з військового обліку"
DIA_BRON        = "Заброньовано"
DIA_NOT_SUBJECT = "Не підлягає бронюванню"
DIA_NOT_BRON    = "Не заброньовано"
DIA_BENEF_SHEET = "бенефіціарн"   # частина назви аркуша для виключення
# ───────────────────────────────────────────────────────────

CLR = {
    "dark_blue": "#1F4E79",
    "mid_blue":  "#2E75B6",
    "green_bg":  "#E2EFDA",
    "green_fg":  "#375623",
    "red_bg":    "#FFE0E0",
    "red_fg":    "#CC0000",
    "grey_fg":   "#595959",
    "white":     "#FFFFFF",
    "black":     "#000000",
    "zebra":     "#F5F9FF",
    "bg":        "#F0F4F8",
    "toolbar":   "#D6E4F0",
    "orange":    "#E67E22",
    "purple":    "#8E44AD",
}

# ══════════════════════════════════════════════════════════
#  ЗБЕРЕЖЕННЯ НАЛАШТУВАНЬ
# ══════════════════════════════════════════════════════════

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, encoding="utf-8") as f:
                data = json.load(f)
            data.setdefault("min_salary",     DEFAULT_MIN_SAL)
            data.setdefault("coefficients",   [dict(c) for c in DEFAULT_COEFFS])
            data.setdefault("bron_limit_pct", 50)
            data.setdefault("recent_files",   [])
            data.setdefault("recent_dia",     [])
            data.setdefault("auto_load_dia",  True)
            return data
        except Exception:
            pass
    return {"min_salary": DEFAULT_MIN_SAL,
            "coefficients": [dict(c) for c in DEFAULT_COEFFS],
            "bron_limit_pct": 50,
            "recent_files": [], "recent_dia": [],
            "auto_load_dia": True}

def save_settings(data):
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# ══════════════════════════════════════════════════════════
#  АВТОВИЗНАЧЕННЯ КОЛОНОК
# ══════════════════════════════════════════════════════════

def detect_columns(headers):
    h_map = {str(h).strip(): h for h in headers if h is not None}
    result = {}
    for role, candidates in CANDIDATE_COLS.items():
        result[role] = next((c for c in candidates if c in h_map), None)
    result["bron"] = next(
        (h for h in headers if h and "бронь" in str(h).lower()), None
    )
    return result

def detect_filter_value(rows, group_col):
    if not group_col:
        return None
    vals = {str(r.get(group_col) or "").strip() for r in rows[:200]}
    return next((c for c in FILTER_CANDIDATES if c in vals), None)

def is_bron_yes(val):
    return str(val).strip().lower() in BRON_YES_VALUES

# ══════════════════════════════════════════════════════════
#  ЧИТАННЯ ФАЙЛУ ДІЇ
# ══════════════════════════════════════════════════════════

def _norm_name(s):
    """Нормалізація ПІБ: прибираємо зайві пробіли, нижній регістр."""
    return re.sub(r"\s+", " ", str(s).strip()).lower()

def read_dia_file(path):
    """Читає xlsx Дії, повертає список dict з ПІБ, Статус, Примітка, Тип.
    Зʼєднує Прізвище+Імʼя+По батькові.
    Пропускає аркуш бенефіціарів. Другий аркуш — тимчасово заброньовані."""
    wb = openpyxl.load_workbook(path, data_only=True)
    all_persons = []
    first_sheet = True
    for sn in wb.sheetnames:
        # Пропускаємо бенефіціарів
        if DIA_BENEF_SHEET in sn.lower():
            continue
        ws = wb[sn]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        h_lower = [str(h).strip().lower() if h else "" for h in headers]
        # Шукаємо колонки
        idx_last  = next((i for i, h in enumerate(h_lower) if "прізвище" in h), None)
        idx_first = next((i for i, h in enumerate(h_lower) if h == "імʼя" or "ім'я" in h or h == "імя" or h.startswith("ім")), None)
        idx_mid   = next((i for i, h in enumerate(h_lower) if "батькові" in h or "по батькові" in h), None)
        idx_stat  = next((i for i, h in enumerate(h_lower) if "статус" in h), None)
        idx_note  = next((i for i, h in enumerate(h_lower) if "примітка" in h), None)
        if idx_last is None or idx_stat is None:
            first_sheet = False
            continue
        # Якщо це не перший аркуш — тимчасово заброньовані
        is_temp = not first_sheet
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[idx_last]:
                continue
            parts = [str(row[idx_last] or "").strip()]
            if idx_first is not None:
                parts.append(str(row[idx_first] or "").strip())
            if idx_mid is not None:
                parts.append(str(row[idx_mid] or "").strip())
            pib = " ".join(p for p in parts if p)
            status = str(row[idx_stat] or "").strip() if idx_stat is not None else ""
            note   = str(row[idx_note] or "").strip() if idx_note is not None else ""
            typ = "тимчасово" if is_temp else "основний"
            all_persons.append({"pib": pib, "pib_key": _norm_name(pib),
                                "status": status, "note": note, "type": typ})
        first_sheet = False
    wb.close()
    return all_persons


def clean_dates_in_sheet(ws):
    """Очищає колонку 'Період' від часу 00:00:00.
    Підтримує два варіанти:
    - текст 'DD.MM.YYYY 00:00:00' → 'DD.MM.YYYY'
    - datetime з часом 00:00:00 → числовий формат DD.MM.YYYY
    Повертає кількість оновлених клітинок."""
    # Знайдемо колонку з періодом/датою
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    date_cols = [i + 1 for i, h in enumerate(headers)
                 if h and ("період" in str(h).lower() or "дата" in str(h).lower())]
    if not date_cols:
        return 0
    count = 0
    for row in ws.iter_rows(min_row=2):
        for col_idx in date_cols:
            if col_idx > len(row):
                continue
            cell = row[col_idx - 1]
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str):
                # Текст із часом — видаляємо " 00:00:00" (або інший час)
                m = re.match(r"^\s*(\d{1,2}\.\d{1,2}\.\d{4})\s+\d{1,2}:\d{2}(:\d{2})?\s*$", v)
                if m:
                    cell.value = m.group(1)
                    count += 1
            elif isinstance(v, datetime):
                # datetime — встановити формат без часу
                if cell.number_format != "DD.MM.YYYY":
                    cell.number_format = "DD.MM.YYYY"
                    count += 1
    return count


def _parse_dia_date(note):
    """Парсить дату з примітки виду 'До 18.03.2027'."""
    m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", note or "")
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    return None

def analyse_dia(dia_data, bron_limit_pct=50):
    """Аналітика по файлу Дії.
    bron_limit_pct — ліміт бронювання у % від військовозобовʼязаних."""
    total      = len(dia_data)
    excluded   = [p for p in dia_data if DIA_EXCLUDED in p.get("note", "")]
    liable     = [p for p in dia_data if DIA_EXCLUDED not in p.get("note", "")]
    bronied    = [p for p in liable if p["status"] == DIA_BRON and p["type"] == "основний"]
    temp_bron  = [p for p in liable if p["status"] == DIA_BRON and p["type"] == "тимчасово"]
    all_bron   = bronied + temp_bron
    deferred   = [p for p in liable if "відстрочк" in p.get("note", "").lower()]
    not_bron   = [p for p in liable
                  if p not in all_bron and p not in deferred]

    n_liable = len(liable)
    n_all_bron = len(all_bron)
    max_allowed = int(n_liable * bron_limit_pct / 100)
    over_limit  = max(0, n_all_bron - max_allowed)
    remaining   = max(0, max_allowed - n_all_bron)
    pct_actual  = (n_all_bron / n_liable * 100) if n_liable else 0
    pct_defer   = (len(deferred) / n_liable * 100) if n_liable else 0
    pct_not_bron = (len(not_bron) / n_liable * 100) if n_liable else 0

    # Аналіз строків бронювання
    now = datetime.now()
    expiring_soon = []   # закінчується протягом 60 днів
    expired = []         # вже прострочено
    dates_found = []
    for p in all_bron:
        dt = _parse_dia_date(p.get("note", ""))
        if dt:
            dates_found.append(dt)
            days_left = (dt - now).days
            if days_left < 0:
                expired.append((p, days_left))
            elif days_left <= 60:
                expiring_soon.append((p, days_left))

    earliest_exp = min(dates_found) if dates_found else None
    latest_exp   = max(dates_found) if dates_found else None

    return {
        "total": total,
        "excluded": len(excluded),
        "liable": n_liable,
        "bronied": len(bronied),
        "temp_bron": len(temp_bron),
        "all_bron": n_all_bron,
        "deferred": len(deferred),
        "not_bron": len(not_bron),
        "bron_limit_pct": bron_limit_pct,
        "max_allowed": max_allowed,
        "over_limit": over_limit,
        "remaining": remaining,
        "pct_bron_of_liable": pct_actual,
        "pct_defer": pct_defer,
        "pct_not_bron": pct_not_bron,
        "expired": expired,
        "expiring_soon": expiring_soon,
        "earliest_exp": earliest_exp,
        "latest_exp": latest_exp,
        "persons": dia_data,
        "liable_persons": liable,
        "bronied_persons": bronied,
        "temp_bron_persons": temp_bron,
        "all_bron_persons": all_bron,
        "deferred_persons": deferred,
        "excluded_persons": excluded,
    }

# ══════════════════════════════════════════════════════════
#  АНАЛІЗ ЗП
# ══════════════════════════════════════════════════════════

def read_sheet(ws):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    return rows, headers

def analyse(rows, headers, min_salary=DEFAULT_MIN_SAL, coeff=2.5):
    threshold  = min_salary * coeff
    cols       = detect_columns(headers)
    filter_val = detect_filter_value(rows, cols["group"])

    emp_col  = cols["employee"] or ""
    dept_col = cols["dept"]
    amt_col  = cols["amount"]  or ""
    grp_col  = cols["group"]
    bron_col = cols["bron"]    or "Бронь"

    if grp_col and filter_val:
        filtered = [r for r in rows
                    if str(r.get(grp_col) or "").strip() == filter_val]
    else:
        filtered = rows

    agg = {}
    for r in filtered:
        name     = str(r.get(emp_col) or "").strip()
        bron_raw = r.get(bron_col)
        if name not in agg:
            agg[name] = {
                "sum":      0.0,
                "dept":     str(r.get(dept_col) or "") if dept_col else "",
                "bron_raw": bron_raw,
            }
        if is_bron_yes(bron_raw):
            agg[name]["bron_raw"] = bron_raw
        try:
            agg[name]["sum"] += float(r.get(amt_col) or 0)
        except (TypeError, ValueError):
            pass

    result = []
    for name, v in sorted(agg.items()):
        suma     = v["sum"]
        bron_raw = v["bron_raw"]
        if is_bron_yes(bron_raw):
            if suma >= threshold:
                status, donar = STATUS_OK, ""
            else:
                status = STATUS_NEED
                donar  = threshold - suma
        else:
            status, donar = "—", ""
        result.append({"name": name, "dept": v["dept"],
                        "bron": bron_raw, "suma": suma,
                        "status": status, "donar": donar})

    n           = len(result)
    total_sum   = sum(r["suma"] for r in result)
    bron_count  = sum(1 for r in result if is_bron_yes(r["bron"]))
    ok_count    = sum(1 for r in result if r["status"] == STATUS_OK)
    need_count  = sum(1 for r in result if r["status"] == STATUS_NEED)
    donar_total = sum(r["donar"] for r in result if r["donar"] != "")
    bron_sum    = sum(r["suma"] for r in result if is_bron_yes(r["bron"]))
    avg_all     = total_sum / n if n else 0
    need_total  = bron_count * threshold
    deficit     = bron_sum - need_total

    return result, dict(
        cols=cols, filter_val=filter_val,
        min_salary=min_salary, coeff=coeff, threshold=threshold,
        bron_count=bron_count, ok_count=ok_count, need_count=need_count,
        donar_total=donar_total, unique_n=n, total_sum=total_sum,
        bron_sum=bron_sum, avg_all=avg_all,
        need_total=need_total, deficit=deficit,
    )

# ══════════════════════════════════════════════════════════
#  ЗАПИС В EXCEL
# ══════════════════════════════════════════════════════════

def _fill(h):  return PatternFill("solid", fgColor=h)
def _font(bold=False, italic=False, color="000000", size=11):
    return Font(bold=bold, italic=italic, color=color, size=size)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def write_excel(wb, data, stats, source_sheet):
    min_salary = stats["min_salary"]
    coeff      = stats["coeff"]
    threshold  = stats["threshold"]

    if OUTPUT_SHEET in wb.sheetnames:
        del wb[OUTPUT_SHEET]
    ws = wb.create_sheet(OUTPUT_SHEET)
    row = 1

    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row, 1, "АНАЛІЗ НАРАХУВАНЬ — ЗАБРОНЬОВАНІ ПРАЦІВНИКИ")
    c.font = _font(bold=True, color="FFFFFF", size=13)
    c.fill = _fill("1F4E79"); c.alignment = _align("center")
    ws.row_dimensions[row].height = 30; row += 1

    for label, val, fmt in [
        ("Мінімальна ЗП в Україні (база):",            min_salary, "#,##0.00"),
        (f"Коефіцієнт для заброньованих (×{coeff}):", coeff,      "0.0#"),
        ("Поріг нарахування для заброньованих:",       threshold,  "#,##0.00"),
    ]:
        ws.cell(row, 1, label).font = _font(bold=True)
        c = ws.cell(row, 2, val)
        c.font = _font(bold=True, color="0000FF"); c.number_format = fmt; row += 1

    cols = stats["cols"]
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row, 1,
        f"Аркуш: '{source_sheet}' | Фільтр: {stats['filter_val'] or '—'} | "
        f"Працівник: {cols['employee'] or '—'} | Сума: {cols['amount'] or '—'} | "
        f"Бронь: {cols['bron'] or '—'} | Згенеровано: {datetime.now():%d.%m.%Y %H:%M}"
    ).font = _font(italic=True, color="595959", size=9)
    row += 2

    for col, t in enumerate(["№", "Співробітник", "Підрозділ", "Бронь",
                              "Сума нараховано", "Статус",
                              f"Донарахувати до {threshold:,.2f}"], 1):
        c = ws.cell(row, col, t)
        c.font = _font(bold=True, color="FFFFFF")
        c.fill = _fill("2E75B6"); c.border = _border()
        c.alignment = _align("center" if col in (1, 4, 6) else "left")
    row += 1

    for i, r in enumerate(data):
        z = _fill("F5F9FF") if i % 2 == 0 else PatternFill()
        ws.cell(row, 1, i + 1).alignment = _align("center")
        ws.cell(row, 2, r["name"]); ws.cell(row, 3, r["dept"])
        cb = ws.cell(row, 4, r["bron"]); cb.alignment = _align("center")
        if is_bron_yes(r["bron"]):
            cb.font = _font(bold=True, color="0070C0")
        cs = ws.cell(row, 5, r["suma"])
        cs.number_format = "#,##0.00"; cs.alignment = _align("right")
        cf = ws.cell(row, 6, r["status"]); cf.alignment = _align("center")
        if r["status"] == STATUS_NEED:
            cf.fill = _fill("FFE0E0"); cf.font = _font(bold=True, color="CC0000")
        elif r["status"] == STATUS_OK:
            cf.fill = _fill("E2EFDA"); cf.font = _font(color="375623")
        if r["donar"] != "":
            cd = ws.cell(row, 7, r["donar"])
            cd.number_format = "#,##0.00"
            cd.font = _font(bold=True, color="CC0000")
            cd.alignment = _align("right")
        for col in range(1, 8):
            c = ws.cell(row, col); c.border = _border()
            if not c.fill or c.fill.fill_type == "none": c.fill = z
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row, 1, "ПІДСУМКИ")
    c.font = _font(bold=True, color="FFFFFF"); c.fill = _fill("1F4E79")
    c.alignment = _align("center"); row += 1

    for label, val, bg, fg in [
        ("Всього заброньованих:",           stats["bron_count"],  None,     None),
        ("Поріг виконано:",                 stats["ok_count"],    "E2EFDA", "375623"),
        ("Потрібно донарахувати (осіб):",   stats["need_count"],  "FFE0E0", "CC0000"),
        ("Загальна сума до донарахування:", stats["donar_total"], None,     "CC0000"),
    ]:
        ws.cell(row, 2, label).font = _font(bold=True)
        c = ws.cell(row, 3, val); c.alignment = _align("center")
        if isinstance(val, float): c.number_format = "#,##0.00"
        if bg: c.fill = _fill(bg)
        if fg: c.font = _font(bold=True, color=fg)
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row, 1,
        f"ЗП ЗАБРОНЬОВАНИХ vs ПОРІГ {threshold:,.2f} грн  (Мін. {min_salary:,} × {coeff})")
    c.font = _font(bold=True, color="FFFFFF"); c.fill = _fill("1F4E79")
    c.alignment = _align("center"); row += 2

    for col, t in enumerate(["Показник", "Значення", "Поріг", "Відповідність"], 2):
        c = ws.cell(row, col, t)
        c.font = _font(bold=True, color="FFFFFF")
        c.fill = _fill("2E75B6"); c.border = _border()
    row += 1

    for j, (label, val, minimum, status) in enumerate([
        ("Загальна сума нарахувань (всі пр-ки)", stats["total_sum"],  "—", "—"),
        ("Кількість заброньованих",              stats["bron_count"], "—", "—"),
        ("Сума нарахувань заброньованих",        stats["bron_sum"],   "—", "—"),
        ("Необхідна загальна сума (поріг × N)",  stats["need_total"], "—", "—"),
        ("Середня ЗП (всі працівники)",             stats["avg_all"],   threshold,
            "✓ ВІДПОВІДАЄ" if stats["avg_all"] >= threshold else "✗ НЕ ВІДПОВІДАЄ"),
        ("Дефіцит / Профіцит заброньованих",    stats["deficit"], "(факт − поріг×N)",
            "✓ ПРОФІЦИТ" if stats["deficit"] >= 0 else "✗ ДЕФІЦИТ"),
    ]):
        z = _fill("F5F9FF") if j % 2 == 0 else PatternFill()
        c_l = ws.cell(row, 2, label); c_l.border = _border()
        if not c_l.fill or c_l.fill.fill_type == "none": c_l.fill = z
        c_v = ws.cell(row, 3, val); c_v.alignment = _align("right"); c_v.border = _border()
        if isinstance(val, float): c_v.number_format = "#,##0.00"
        if not c_v.fill or c_v.fill.fill_type == "none": c_v.fill = z
        c_m = ws.cell(row, 4, minimum); c_m.border = _border()
        if isinstance(minimum, (int, float)):
            c_m.number_format = "#,##0.00"; c_m.font = _font(color="0000FF")
        else:
            c_m.font = _font(italic=True, color="595959", size=9)
        c_m.alignment = _align("right")
        if not c_m.fill or c_m.fill.fill_type == "none": c_m.fill = z
        c_s = ws.cell(row, 5, status); c_s.alignment = _align("center"); c_s.border = _border()
        if isinstance(status, str) and ("НЕ" in status or "ДЕФІЦИТ" in status):
            c_s.fill = _fill("FFE0E0"); c_s.font = _font(bold=True, color="CC0000")
            if isinstance(val, float): c_v.font = _font(bold=True, color="CC0000")
        elif isinstance(status, str) and ("ВІДПОВІДАЄ" in status or "ПРОФІЦИТ" in status):
            c_s.fill = _fill("E2EFDA"); c_s.font = _font(bold=True, color="375623")
            if isinstance(val, float): c_v.font = _font(bold=True, color="375623")
        else:
            if not c_s.fill or c_s.fill.fill_type == "none": c_s.fill = z
        row += 1

    for col, w in zip("ABCDEFG", [5, 32, 18, 12, 16, 26, 22]):
        ws.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════
#  АВТОПОШУК xlsx
# ══════════════════════════════════════════════════════════

def find_xlsx():
    here = os.path.dirname(os.path.abspath(__file__))
    files = [f for f in glob.glob(os.path.join(here, "*.xlsx"))
             if not os.path.basename(f).startswith("~$")]
    return files

# ══════════════════════════════════════════════════════════
#  ДІАЛОГ НАЛАШТУВАНЬ КОЕФІЦІЄНТІВ
# ══════════════════════════════════════════════════════════

class SettingsDialog(tk.Toplevel):
    def __init__(self, parent, coeffs, on_save):
        super().__init__(parent)
        self.title("Налаштування коефіцієнтів")
        self.geometry("430x370")
        self.resizable(False, False)
        self.grab_set()
        self.coeffs  = [dict(c) for c in coeffs]
        self.on_save = on_save
        self._build()
        self._refresh()

    def _build(self):
        tk.Label(self,
                 text="Поріг = Мін. ЗП × Коефіцієнт. Можна додавати свої категорії.",
                 font=("Segoe UI", 9), fg=CLR["grey_fg"],
                 wraplength=390, justify="left").pack(pady=(10, 4), padx=14, anchor="w")

        lf = tk.Frame(self); lf.pack(fill="x", padx=14)
        self.lb = tk.Listbox(lf, font=("Segoe UI", 11), height=5,
                             selectmode="single", activestyle="none",
                             selectbackground=CLR["mid_blue"], selectforeground="white")
        self.lb.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(lf, orient="vertical", command=self.lb.yview)
        sb.pack(side="right", fill="y")
        self.lb.configure(yscrollcommand=sb.set)
        self.lb.bind("<<ListboxSelect>>", self._on_select)

        ef = tk.LabelFrame(self, text=" Редагування ", font=("Segoe UI", 9))
        ef.pack(fill="x", padx=14, pady=(8, 0))

        tk.Label(ef, text="Назва:", font=("Segoe UI", 9)).grid(
            row=0, column=0, sticky="w", padx=8, pady=6)
        self.name_var = tk.StringVar()
        tk.Entry(ef, textvariable=self.name_var, font=("Segoe UI", 10),
                 width=28).grid(row=0, column=1, sticky="w", padx=8, pady=6)

        tk.Label(ef, text="Коефіцієнт (×):", font=("Segoe UI", 9)).grid(
            row=1, column=0, sticky="w", padx=8, pady=4)
        self.val_var = tk.StringVar()
        tk.Entry(ef, textvariable=self.val_var, font=("Segoe UI", 10),
                 width=10).grid(row=1, column=1, sticky="w", padx=8, pady=4)

        bf = tk.Frame(ef); bf.grid(row=2, column=0, columnspan=2, pady=8)
        for text, cmd, color in [
            ("Додати",   self._add,    CLR["mid_blue"]),
            ("Оновити",  self._update, CLR["mid_blue"]),
            ("Видалити", self._delete, CLR["red_fg"]),
        ]:
            tk.Button(bf, text=text, command=cmd, bg=color, fg="white",
                      font=("Segoe UI", 9), relief="flat",
                      padx=10, pady=3, cursor="hand2").pack(side="left", padx=4)

        bf2 = tk.Frame(self); bf2.pack(pady=12)
        tk.Button(bf2, text="✔  Зберегти і закрити", command=self._save,
                  bg="#27AE60", fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=14, pady=5, cursor="hand2").pack(side="left", padx=6)
        tk.Button(bf2, text="Скасувати", command=self.destroy,
                  bg=CLR["grey_fg"], fg="white", font=("Segoe UI", 10),
                  relief="flat", padx=14, pady=5, cursor="hand2").pack(side="left", padx=6)

    def _refresh(self):
        self.lb.delete(0, "end")
        for c in self.coeffs:
            self.lb.insert("end", f"  {c['name']}   (×{c['value']})")

    def _on_select(self, _=None):
        idx = self.lb.curselection()
        if idx:
            c = self.coeffs[idx[0]]
            self.name_var.set(c["name"])
            self.val_var.set(str(c["value"]))

    def _parse_val(self):
        try:
            v = float(self.val_var.get().replace(",", "."))
            if v <= 0: raise ValueError
            return v
        except ValueError:
            messagebox.showerror("Помилка", "Коефіцієнт має бути числом > 0.", parent=self)
            return None

    def _add(self):
        name = self.name_var.get().strip()
        if not name:
            messagebox.showerror("Помилка", "Введіть назву.", parent=self); return
        val = self._parse_val()
        if val is None: return
        self.coeffs.append({"name": name, "value": val})
        self._refresh()
        self.lb.selection_set("end")

    def _update(self):
        idx = self.lb.curselection()
        if not idx:
            messagebox.showinfo("", "Оберіть запис для редагування.", parent=self); return
        name = self.name_var.get().strip()
        if not name:
            messagebox.showerror("Помилка", "Введіть назву.", parent=self); return
        val = self._parse_val()
        if val is None: return
        self.coeffs[idx[0]] = {"name": name, "value": val}
        self._refresh()
        self.lb.selection_set(idx[0])

    def _delete(self):
        idx = self.lb.curselection()
        if not idx:
            messagebox.showinfo("", "Оберіть запис для видалення.", parent=self); return
        if len(self.coeffs) <= 1:
            messagebox.showerror("Помилка", "Потрібен хоча б один коефіцієнт.", parent=self)
            return
        self.coeffs.pop(idx[0])
        self._refresh()

    def _save(self):
        self.on_save(self.coeffs)
        self.destroy()

# ══════════════════════════════════════════════════════════
#  GUI — ГОЛОВНЕ ВІКНО
# ══════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Аналіз бронь — ЗП")
        self.geometry("1200x840")
        self.minsize(900, 640)
        self.configure(bg=CLR["bg"])
        self.settings      = load_settings()
        self.current_file  = None
        self.current_sheet = None
        self.data, self.stats, self._wb = [], {}, None
        self.dia_data  = None   # дані з файлу Дії
        self.dia_stats = None
        self._sort_states = {}  # tree_id -> {col: ascending}
        self._bron_limit_var = tk.IntVar(value=self.settings.get("bron_limit_pct", 50))
        self._min_sal_var = tk.IntVar(value=self.settings["min_salary"])
        self._coeff_var   = tk.StringVar()
        self._build_ui()
        self.after(100, self._auto_open)

    # ── Автовідкриття ─────────────────────────────────────

    def _auto_open(self):
        files = find_xlsx()
        if len(files) == 1:
            self._load_file(files[0])
        elif len(files) > 1:
            self._show_file_picker(files)
        # Авто-завантаження файлу Дії
        if self.settings.get("auto_load_dia", True):
            for path in self.settings.get("recent_dia", []):
                if os.path.exists(path):
                    self.after(300, lambda p=path: self._load_dia_file_silent(p))
                    break

    def _show_file_picker(self, files):
        win = tk.Toplevel(self)
        win.title("Оберіть файл"); win.geometry("520x300"); win.grab_set()
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        tk.Label(win, text="Знайдено кілька xlsx. Оберіть потрібний:",
                 font=("Segoe UI", 10)).pack(pady=12, padx=12)
        lb = tk.Listbox(win, font=("Segoe UI", 10))
        lb.pack(fill="both", expand=True, padx=12)
        for f in files: lb.insert("end", os.path.basename(f))
        lb.select_set(0)
        def ok():
            idx = lb.curselection()
            if idx:
                path = files[idx[0]]
                win.destroy()
                self._load_file(path)
        lb.bind("<Double-1>", lambda _: ok())
        tk.Button(win, text="Відкрити", command=ok,
                  bg=CLR["mid_blue"], fg="white",
                  font=("Segoe UI", 10, "bold"), relief="flat",
                  padx=16, pady=4).pack(pady=10)

    def _pick_sheet(self, sheetnames, callback):
        candidates = [s for s in sheetnames if s != OUTPUT_SHEET]
        if not candidates:
            messagebox.showerror("Помилка", "Не знайдено придатного аркуша."); return
        if len(candidates) == 1:
            callback(candidates[0]); return
        win = tk.Toplevel(self)
        win.title("Оберіть аркуш"); win.geometry("400x280"); win.grab_set()
        win.protocol("WM_DELETE_WINDOW", win.destroy)
        tk.Label(win, text="Оберіть аркуш з даними:",
                 font=("Segoe UI", 10, "bold")).pack(pady=14, padx=14)
        lb = tk.Listbox(win, font=("Segoe UI", 11), height=8)
        lb.pack(fill="both", expand=True, padx=14)
        for s in candidates: lb.insert("end", s)
        lb.select_set(0)
        def ok():
            idx = lb.curselection()
            if idx:
                sheet = candidates[idx[0]]
                win.destroy()
                callback(sheet)
        lb.bind("<Double-1>", lambda _: ok())
        tk.Button(win, text="Обрати", command=ok,
                  bg=CLR["mid_blue"], fg="white",
                  font=("Segoe UI", 10, "bold"), relief="flat",
                  padx=16, pady=4).pack(pady=10)

    # ── Побудова інтерфейсу ───────────────────────────────

    def _build_ui(self):
        # ── Меню ──────────────────────────────────────────
        self._build_menu()

        # Верхня панель
        top = tk.Frame(self, bg=CLR["dark_blue"], height=54)
        top.pack(fill="x"); top.pack_propagate(False)
        tk.Label(top, text="📋  Аналіз нарахувань — заброньовані працівники",
                 bg=CLR["dark_blue"], fg=CLR["white"],
                 font=("Segoe UI", 13, "bold")).pack(side="left", padx=16, pady=10)
        bf = tk.Frame(top, bg=CLR["dark_blue"]); bf.pack(side="right", padx=12)
        self.btn_save = tk.Button(bf, text="💾  Зберегти в Excel",
            command=self._save_excel, state="disabled",
            bg="#27AE60", fg="white", font=("Segoe UI", 10, "bold"),
            relief="flat", padx=12, pady=4, cursor="hand2")
        self.btn_save.pack(side="right", padx=4)
        tk.Button(bf, text="📂  Відкрити файл", command=self._open_file,
            bg=CLR["mid_blue"], fg="white", font=("Segoe UI", 10, "bold"),
            relief="flat", padx=12, pady=4, cursor="hand2").pack(side="right", padx=4)
        tk.Button(bf, text="🪖  Завантажити Дію", command=self._open_dia_file,
            bg=CLR["orange"], fg="white", font=("Segoe UI", 10, "bold"),
            relief="flat", padx=12, pady=4, cursor="hand2").pack(side="right", padx=4)

        # Панель налаштувань
        tb = tk.Frame(self, bg=CLR["toolbar"], height=42)
        tb.pack(fill="x"); tb.pack_propagate(False)

        tk.Label(tb, text="Мін. ЗП:", bg=CLR["toolbar"],
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=(12, 2), pady=8)
        vcmd = self.register(lambda v: v.isdigit() or v == "")
        tk.Spinbox(tb, from_=1, to=999_999, increment=100,
                   textvariable=self._min_sal_var,
                   font=("Segoe UI", 10), width=8,
                   validate="key", validatecommand=(vcmd, "%P")
                   ).pack(side="left", pady=6)
        self._min_sal_var.trace("w", lambda *_: self._update_threshold_label())
        tk.Label(tb, text="грн", bg=CLR["toolbar"],
                 font=("Segoe UI", 9)).pack(side="left", padx=(2, 10))

        tk.Label(tb, text="Коефіцієнт:", bg=CLR["toolbar"],
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 4))
        self._coeff_cb = ttk.Combobox(tb, textvariable=self._coeff_var,
                                      state="readonly", width=18,
                                      font=("Segoe UI", 10))
        self._coeff_cb.pack(side="left", pady=6)
        self._coeff_cb.bind("<<ComboboxSelected>>",
                            lambda *_: self._update_threshold_label())

        tk.Label(tb, text="= Поріг:", bg=CLR["toolbar"],
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=(10, 4))
        self._threshold_lbl_var = tk.StringVar(value="—")
        tk.Label(tb, textvariable=self._threshold_lbl_var,
                 bg=CLR["toolbar"], fg=CLR["dark_blue"],
                 font=("Segoe UI", 11, "bold"), width=12).pack(side="left")
        tk.Label(tb, text="грн", bg=CLR["toolbar"],
                 font=("Segoe UI", 9)).pack(side="left", padx=(2, 6))

        tk.Button(tb, text="↺ Перерахувати", command=self._reanalyse,
                  bg=CLR["mid_blue"], fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=8, pady=2, cursor="hand2").pack(side="left", padx=2)
        tk.Button(tb, text="⚙", command=self._open_settings,
                  bg=CLR["dark_blue"], fg="white", font=("Segoe UI", 11),
                  relief="flat", padx=6, pady=2, cursor="hand2").pack(side="left", padx=6)

        tk.Label(tb, text="│", bg=CLR["toolbar"],
                 fg="#9ABBE0", font=("Segoe UI", 14)).pack(side="left", padx=8)

        tk.Label(tb, text="Аркуш:", bg=CLR["toolbar"],
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 4))
        self._sheet_var = tk.StringVar()
        self._sheet_cb  = ttk.Combobox(tb, textvariable=self._sheet_var,
                                       state="disabled", width=24,
                                       font=("Segoe UI", 10))
        self._sheet_cb.pack(side="left", pady=6)
        self._sheet_cb.bind("<<ComboboxSelected>>", self._on_sheet_change)

        tk.Label(tb, text="│", bg=CLR["toolbar"],
                 fg="#9ABBE0", font=("Segoe UI", 14)).pack(side="left", padx=8)

        tk.Label(tb, text="Ліміт броні:", bg=CLR["toolbar"],
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 2))
        tk.Spinbox(tb, from_=1, to=100, increment=5,
                   textvariable=self._bron_limit_var,
                   font=("Segoe UI", 9), width=4,
                   command=self._on_bron_limit_change).pack(side="left", pady=6)
        self._bron_limit_var.trace("w", lambda *_: self._on_bron_limit_change())
        tk.Label(tb, text="%", bg=CLR["toolbar"],
                 font=("Segoe UI", 9)).pack(side="left", padx=(1, 6))

        # Статусний рядок
        self.status_var = tk.StringVar(value="Шукаю xlsx у поточній папці…")
        tk.Label(self, textvariable=self.status_var,
                 bg=CLR["bg"], fg=CLR["grey_fg"],
                 font=("Segoe UI", 9, "italic"), anchor="w"
                 ).pack(fill="x", padx=14, pady=(4, 0))

        # Вкладки
        style = ttk.Style(self); style.theme_use("clam")
        style.configure("TNotebook", background=CLR["bg"], borderwidth=0)
        style.configure("TNotebook.Tab", font=("Segoe UI", 10), padding=[14, 6])
        style.map("TNotebook.Tab",
                  background=[("selected", CLR["mid_blue"]), ("!selected", "#CBD8E8")],
                  foreground=[("selected", "white"), ("!selected", CLR["dark_blue"])])

        # Банер попереджень (поза вкладками)
        self._dia_warn_var = tk.StringVar(value="")
        self._dia_warn_frame = tk.Frame(self, bg=CLR["red_bg"])
        tk.Label(self._dia_warn_frame, textvariable=self._dia_warn_var,
                 bg=CLR["red_bg"], fg=CLR["red_fg"],
                 font=("Segoe UI", 9, "bold"), anchor="w",
                 wraplength=1100, justify="left").pack(fill="x", padx=10, pady=4)

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=10, pady=6)
        self.tab_main    = tk.Frame(self.nb, bg=CLR["bg"])
        self.tab_summary = tk.Frame(self.nb, bg=CLR["bg"])
        self.tab_avg     = tk.Frame(self.nb, bg=CLR["bg"])
        self.nb.add(self.tab_main,    text="  📊 Зведена таблиця  ")
        self.nb.add(self.tab_summary, text="  📌 Підсумки  ")
        self.nb.add(self.tab_avg,     text="  📈 Аналітика  ")
        self._build_main_tab()
        self._build_summary_tab()
        self._build_avg_tab()

        # Ініціалізація коефіцієнтів і порогу
        self._refresh_coeff_cb()

    # ── Коефіцієнти ───────────────────────────────────────

    def _refresh_coeff_cb(self):
        names = [c["name"] for c in self.settings["coefficients"]]
        self._coeff_cb.configure(values=names)
        if names and self._coeff_var.get() not in names:
            self._coeff_var.set(names[0])
        self._update_threshold_label()

    def _get_coeff(self):
        name = self._coeff_var.get()
        for c in self.settings["coefficients"]:
            if c["name"] == name:
                return c["value"]
        return self.settings["coefficients"][0]["value"] if self.settings["coefficients"] else 2.5

    def _get_min_sal(self):
        try:
            return max(1, int(self._min_sal_var.get()))
        except (ValueError, tk.TclError):
            return DEFAULT_MIN_SAL

    def _update_threshold_label(self):
        try:
            self._threshold_lbl_var.set(
                f"{self._get_min_sal() * self._get_coeff():,.2f}")
        except Exception:
            self._threshold_lbl_var.set("—")

    def _open_settings(self):
        SettingsDialog(self, self.settings["coefficients"], self._on_settings_saved)

    def _on_settings_saved(self, coeffs):
        self.settings["coefficients"] = coeffs
        self.settings["min_salary"]   = self._get_min_sal()
        save_settings(self.settings)
        self._refresh_coeff_cb()

    # ── Картки ────────────────────────────────────────────

    def _card(self, parent, label, value, color, wide=False):
        f = tk.Frame(parent, bg=color, width=175 if wide else 138, height=64)
        f.pack(side="left", padx=5); f.pack_propagate(False)
        tk.Label(f, text=label, bg=color, fg="white", font=("Segoe UI", 8)).pack(pady=(8, 0))
        lbl = tk.Label(f, text=value, bg=color, fg="white", font=("Segoe UI", 14, "bold"))
        lbl.pack(); return lbl

    # ── Сортування Treeview ──────────────────────────────

    def _sort_tree(self, tree, col):
        tid = id(tree)
        if tid not in self._sort_states:
            self._sort_states[tid] = {}
        ascending = not self._sort_states[tid].get(col, False)
        self._sort_states[tid][col] = ascending

        items = [(tree.set(k, col), k) for k in tree.get_children("")]

        def sort_key(item):
            val = item[0]
            # Спроба числового сортування
            try:
                return (0, float(val.replace(",", "").replace(" ", "")
                                    .replace("грн", "").replace("+", "").strip()))
            except (ValueError, AttributeError):
                return (1, val.lower())

        items.sort(key=sort_key, reverse=not ascending)
        for idx, (_, k) in enumerate(items):
            tree.move(k, "", idx)

        # Оновити заголовок зі стрілкою
        for c in tree["columns"]:
            base = c.split(" ▲")[0].split(" ▼")[0]
            arrow = ""
            if c == col:
                arrow = " ▲" if ascending else " ▼"
            tree.heading(c, text=base + arrow)

    def _make_tree(self, parent, cols, widths, height=18):
        style = ttk.Style(); tid = "T.Treeview"
        style.configure(tid, background="white", foreground=CLR["black"],
                        rowheight=26, font=("Segoe UI", 10),
                        fieldbackground="white", borderwidth=0)
        style.configure(f"{tid}.Heading", background=CLR["mid_blue"], foreground="white",
                        font=("Segoe UI", 10, "bold"), relief="flat")
        style.map(tid,
                  background=[("selected", CLR["mid_blue"])],
                  foreground=[("selected", "white")])
        frame = tk.Frame(parent, bg=CLR["bg"])
        frame.pack(fill="both", expand=True, padx=10, pady=6)
        tree = ttk.Treeview(frame, columns=cols, show="headings", height=height, style=tid)
        for col, w in zip(cols, widths):
            tree.heading(col, text=col,
                         command=lambda c=col, t=tree: self._sort_tree(t, c))
            tree.column(col, width=w, minwidth=40,
                        anchor="center" if col in ("№", "Бронь", "Відповідність") else "w")
        vsb = ttk.Scrollbar(frame, orient="vertical",   command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1); frame.columnconfigure(0, weight=1)
        tree.tag_configure("need",    background=CLR["red_bg"],   foreground=CLR["red_fg"])
        tree.tag_configure("ok",      background=CLR["green_bg"], foreground=CLR["green_fg"])
        tree.tag_configure("zebra",   background=CLR["zebra"])
        tree.tag_configure("deficit", background=CLR["red_bg"],   foreground=CLR["red_fg"])
        tree.tag_configure("surplus", background=CLR["green_bg"], foreground=CLR["green_fg"])
        tree.tag_configure("red",     foreground=CLR["red_fg"])
        tree.tag_configure("bron",    foreground="#0070C0")
        tree.tag_configure("defer",   foreground=CLR["orange"])
        tree.tag_configure("excl",    foreground=CLR["grey_fg"])
        tree.tag_configure("expired", background="#FFE0E0", foreground=CLR["red_fg"])
        tree.tag_configure("expiring", background="#FFF2CC", foreground="#806000")
        return tree

    # ── Меню та гарячі клавіші ────────────────────────────

    def _build_menu(self):
        menubar = tk.Menu(self)
        # File
        m_file = tk.Menu(menubar, tearoff=0)
        m_file.add_command(label="Відкрити xlsx ЗП…   Ctrl+O",
                            command=self._open_file)
        m_file.add_command(label="Відкрити файл Дії…  Ctrl+D",
                            command=self._open_dia_file)
        m_file.add_separator()
        self._recent_menu = tk.Menu(m_file, tearoff=0)
        m_file.add_cascade(label="Нещодавні файли ЗП", menu=self._recent_menu)
        self._recent_dia_menu = tk.Menu(m_file, tearoff=0)
        m_file.add_cascade(label="Нещодавні файли Дії", menu=self._recent_dia_menu)
        m_file.add_separator()
        m_file.add_command(label="Зберегти в Excel    Ctrl+S",
                            command=self._save_excel)
        m_file.add_command(label="Експорт CSV…        Ctrl+E",
                            command=lambda: self._export_csv(self.tree))
        m_file.add_separator()
        m_file.add_command(label="Вийти               Ctrl+Q",
                            command=self.destroy)
        menubar.add_cascade(label="Файл", menu=m_file)

        # Settings
        m_set = tk.Menu(menubar, tearoff=0)
        m_set.add_command(label="Коефіцієнти…", command=self._open_settings)
        self._auto_load_var = tk.BooleanVar(value=self.settings.get("auto_load_dia", True))
        m_set.add_checkbutton(label="Авто-завантажувати останній файл Дії",
                               variable=self._auto_load_var,
                               command=self._on_auto_load_toggle)
        menubar.add_cascade(label="Налаштування", menu=m_set)

        # Help
        m_help = tk.Menu(menubar, tearoff=0)
        m_help.add_command(label="Легенда кольорів", command=self._show_legend)
        m_help.add_command(label="Гарячі клавіші", command=self._show_hotkeys_help)
        m_help.add_command(label="Про програму", command=self._show_about)
        menubar.add_cascade(label="Довідка", menu=m_help)

        self.config(menu=menubar)
        self._refresh_recent_menu()

        # Hotkeys
        self.bind_all("<Control-o>", lambda e: self._open_file())
        self.bind_all("<Control-O>", lambda e: self._open_file())
        self.bind_all("<Control-d>", lambda e: self._open_dia_file())
        self.bind_all("<Control-D>", lambda e: self._open_dia_file())
        self.bind_all("<Control-s>", lambda e: self._save_excel())
        self.bind_all("<Control-S>", lambda e: self._save_excel())
        self.bind_all("<Control-e>", lambda e: self._export_csv(self.tree))
        self.bind_all("<Control-E>", lambda e: self._export_csv(self.tree))
        self.bind_all("<Control-f>", lambda e: self._focus_search())
        self.bind_all("<Control-F>", lambda e: self._focus_search())
        self.bind_all("<F5>", lambda e: self._reanalyse())
        self.bind_all("<Control-q>", lambda e: self.destroy())
        self.bind_all("<Control-Q>", lambda e: self.destroy())

    def _focus_search(self):
        try:
            self.nb.select(self.tab_main)
            self.search_var.set("")
            # знайдемо поле пошуку
            self.search_entry.focus_set()
        except Exception:
            pass

    def _on_auto_load_toggle(self):
        self.settings["auto_load_dia"] = self._auto_load_var.get()
        save_settings(self.settings)

    def _show_hotkeys_help(self):
        text = (
            "Гарячі клавіші:\n\n"
            "Ctrl+O  — Відкрити файл ЗП\n"
            "Ctrl+D  — Відкрити файл Дії\n"
            "Ctrl+S  — Зберегти в Excel\n"
            "Ctrl+E  — Експорт CSV\n"
            "Ctrl+F  — Пошук у таблиці\n"
            "F5      — Перерахувати\n"
            "Ctrl+Q  — Вийти\n\n"
            "Подвійний клік на рядку — деталі працівника\n"
            "Правий клік на рядку    — копіювати у буфер"
        )
        messagebox.showinfo("Гарячі клавіші", text)

    def _show_about(self):
        messagebox.showinfo("Про програму",
            "Аналіз броньювання — ЗП  (v4)\n\n"
            "Десктопна програма для аналізу зарплати\n"
            "заброньованих працівників та військового обліку.\n\n"
            "GitHub: github.com/r0665629590-droid/bron-analysis")

    def _show_legend(self):
        win = tk.Toplevel(self)
        win.title("Легенда кольорів")
        win.geometry("420x380"); win.resizable(False, False); win.grab_set()
        tk.Label(win, text="Кольори у таблицях", bg=CLR["dark_blue"],
                 fg="white", font=("Segoe UI", 11, "bold"),
                 anchor="w").pack(fill="x", padx=10, pady=(10, 4), ipady=6)
        items = [
            (CLR["green_bg"],  CLR["green_fg"], "OK — поріг ЗП виконано"),
            (CLR["red_bg"],    CLR["red_fg"],   "Потрібно донарахувати"),
            (CLR["zebra"],     CLR["black"],    "Звичайний рядок (зебра)"),
            ("#FFFFFF",        "#0070C0",       "Заброньований працівник"),
            ("#FFFFFF",        CLR["orange"],   "Тимчасова бронь / Відстрочка"),
            ("#FFFFFF",        CLR["grey_fg"],  "Виключено з військового обліку"),
            ("#FFE0E0",        CLR["red_fg"],   "Прострочена бронь (термін минув)"),
            ("#FFF2CC",        "#806000",       "Бронь закінчується < 60 днів"),
        ]
        for bg, fg, text in items:
            f = tk.Frame(win, bg=CLR["bg"])
            f.pack(fill="x", padx=14, pady=3)
            sw = tk.Label(f, text="     ", bg=bg, fg=fg, font=("Segoe UI", 10),
                          relief="solid", borderwidth=1, width=4)
            sw.pack(side="left", padx=(0, 10))
            tk.Label(f, text=text, bg=CLR["bg"], fg=fg,
                     font=("Segoe UI", 10, "bold"), anchor="w").pack(side="left")
        tk.Button(win, text="Закрити", command=win.destroy,
                  bg=CLR["mid_blue"], fg="white", font=("Segoe UI", 10),
                  relief="flat", padx=20, pady=4).pack(pady=12)

    # ── Recent files ─────────────────────────────────────

    def _refresh_recent_menu(self):
        for menu, key, cmd in [(self._recent_menu, "recent_files", self._load_file),
                                (self._recent_dia_menu, "recent_dia", self._load_dia_file_silent)]:
            menu.delete(0, "end")
            files = [f for f in self.settings.get(key, []) if os.path.exists(f)]
            if not files:
                menu.add_command(label="(порожньо)", state="disabled")
            else:
                for f in files[:10]:
                    name = os.path.basename(f)
                    if len(name) > 50:
                        name = name[:47] + "..."
                    menu.add_command(label=name, command=lambda p=f, c=cmd: c(p))

    def _add_recent(self, path, key):
        files = [f for f in self.settings.get(key, []) if f != path and os.path.exists(f)]
        files.insert(0, path)
        self.settings[key] = files[:10]
        save_settings(self.settings)
        self._refresh_recent_menu()

    def _load_dia_file_silent(self, path):
        """Завантаження файлу Дії без діалогу (для recent menu/автозавантаження)."""
        try:
            self.dia_data = read_dia_file(path)
            if not self.dia_data:
                return
            pct = self._bron_limit_var.get()
            self.dia_stats = analyse_dia(self.dia_data, bron_limit_pct=pct)
            self._populate_dia()
            self._update_filter_options()
            self._add_recent(path, "recent_dia")
            if self.data:
                self._populate()
            else:
                self._apply_filters()
                for item in self.tree_avg.get_children(): self.tree_avg.delete(item)
                for label, val, minimum, status, tag in self._build_dia_analytics_rows():
                    self.tree_avg.insert("", "end",
                        values=(label, val, minimum, status), tags=(tag,))
            self.status_var.set(
                f"📂 Дія: {os.path.basename(path)}  │  "
                f"{self.dia_stats['total']} осіб  │  "
                f"{self.dia_stats['liable']} в/з")
        except Exception as e:
            messagebox.showerror("Помилка", f"Не вдалося прочитати файл Дії:\n{e}")

    # ── Контекстне меню та подвійний клік ────────────────

    def _setup_tree_interactions(self, tree, on_double=None):
        """Підключає правий клік (контекст) і подвійний клік до Treeview."""
        menu = tk.Menu(tree, tearoff=0)
        menu.add_command(label="Копіювати клітинку",
                          command=lambda: self._copy_cell(tree))
        menu.add_command(label="Копіювати рядок (з табуляцією)",
                          command=lambda: self._copy_row(tree))
        menu.add_command(label="Копіювати всі видимі рядки",
                          command=lambda: self._copy_all_visible(tree))

        def on_right_click(event):
            row = tree.identify_row(event.y)
            col = tree.identify_column(event.x)
            if row:
                tree.selection_set(row)
                tree._last_col = col
            menu.tk_popup(event.x_root, event.y_root)
        tree.bind("<Button-3>", on_right_click)
        if on_double:
            tree.bind("<Double-1>", lambda e: on_double(tree, e))

    def _copy_cell(self, tree):
        sel = tree.selection()
        if not sel:
            return
        col = getattr(tree, "_last_col", "#1")
        try:
            col_idx = int(col.replace("#", "")) - 1
            val = tree.item(sel[0], "values")[col_idx]
        except Exception:
            val = " ".join(str(v) for v in tree.item(sel[0], "values"))
        self.clipboard_clear(); self.clipboard_append(str(val))

    def _copy_row(self, tree):
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0], "values")
        self.clipboard_clear()
        self.clipboard_append("\t".join(str(v) for v in vals))

    def _copy_all_visible(self, tree):
        cols = tree["columns"]
        lines = ["\t".join(cols)]
        for item in tree.get_children():
            lines.append("\t".join(str(v) for v in tree.item(item, "values")))
        self.clipboard_clear()
        self.clipboard_append("\n".join(lines))
        messagebox.showinfo("Готово", f"Скопійовано {len(lines) - 1} рядків.")

    def _on_main_row_double(self, tree, _evt):
        """Деталі по працівнику — всі вхідні рядки з регістра ЗП."""
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0], "values")
        if len(vals) < 2:
            return
        name = vals[1]
        self._show_employee_details(name)

    def _show_employee_details(self, name):
        if not self._wb or not self.current_sheet:
            messagebox.showinfo("Деталі", f"Працівник: {name}\n\n"
                                  "Завантажте файл ЗП для перегляду деталей.")
            return
        rows, headers = read_sheet(self._wb[self.current_sheet])
        cols_info = detect_columns(headers)
        emp_col = cols_info["employee"] or ""
        amt_col = cols_info["amount"] or ""
        grp_col = cols_info["group"] or ""
        bron_col = cols_info["bron"] or "Бронь"
        # Знайдемо всі рядки цього працівника
        matched = [r for r in rows
                   if _norm_name(str(r.get(emp_col) or "")) == _norm_name(name)]
        win = tk.Toplevel(self)
        win.title(f"Деталі — {name}")
        win.geometry("780x460"); win.grab_set()
        tk.Label(win, text=f"📋  {name}", bg=CLR["dark_blue"], fg="white",
                 font=("Segoe UI", 12, "bold"), anchor="w").pack(
                     fill="x", padx=10, pady=(10, 0), ipady=6)
        tk.Label(win, text=f"Знайдено рядків: {len(matched)}  │  "
                  f"Аркуш: «{self.current_sheet}»  │  Файл: {os.path.basename(self.current_file)}",
                  bg=CLR["bg"], fg=CLR["grey_fg"], font=("Segoe UI", 9, "italic"),
                  anchor="w").pack(fill="x", padx=12, pady=2)
        cols = ("№", "Сума", "Група", "Бронь", "Реєстратор")
        frame = tk.Frame(win, bg=CLR["bg"])
        frame.pack(fill="both", expand=True, padx=10, pady=4)
        tree = ttk.Treeview(frame, columns=cols, show="headings", height=14)
        for c, w in zip(cols, [40, 110, 200, 60, 350]):
            tree.heading(c, text=c)
            tree.column(c, width=w, anchor="w")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        total = 0.0
        for i, r in enumerate(matched, 1):
            try:
                amt = float(r.get(amt_col) or 0)
                total += amt
            except (TypeError, ValueError):
                amt = 0
            reg = ""
            for hdr in headers:
                if hdr and "реєстратор" in str(hdr).lower():
                    reg = str(r.get(hdr) or "")
                    break
            tree.insert("", "end", values=(i, f"{amt:,.2f}",
                                            str(r.get(grp_col) or ""),
                                            str(r.get(bron_col) or ""),
                                            reg))
        tk.Label(win, text=f"Загальна сума: {total:,.2f} грн",
                  bg=CLR["bg"], fg=CLR["dark_blue"],
                  font=("Segoe UI", 11, "bold")).pack(pady=6, anchor="e", padx=14)
        tk.Button(win, text="Закрити", command=win.destroy,
                  bg=CLR["mid_blue"], fg="white", font=("Segoe UI", 10),
                  relief="flat", padx=20, pady=4).pack(pady=(0, 10))

    # ── Вкладки ───────────────────────────────────────────

    def _build_main_tab(self):
        f = self.tab_main
        cf = tk.Frame(f, bg=CLR["bg"]); cf.pack(fill="x", padx=10, pady=(10, 6))
        self.card_total = self._card(cf, "Всього пр-ків",      "—", CLR["mid_blue"])
        self.card_bron  = self._card(cf, "Заброньованих",      "—", "#2980B9")
        self.card_ok    = self._card(cf, "Поріг OK",           "—", CLR["green_fg"])
        self.card_need  = self._card(cf, "Донарахувати",       "—", CLR["red_fg"])
        self.card_donar = self._card(cf, "Сума донарахування", "—", CLR["red_fg"], wide=True)
        cols = ("№", "Співробітник", "Підрозділ", "Бронь",
                "Сума нараховано", "Статус", "Донарахувати",
                "Дія: статус", "Тип броні", "Термін")
        self.tree = self._make_tree(f, cols,
            [40, 200, 110, 60, 110, 160, 100, 165, 90, 100])

        # Фільтри по стовбцях
        ff = tk.Frame(f, bg=CLR["bg"]); ff.pack(fill="x", padx=10, pady=(0, 2))
        tk.Label(ff, text="Фільтри:", bg=CLR["bg"],
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 6))

        self._filter_vars = {}
        for col_name, width in [("Бронь", 6), ("Статус", 18), ("Підрозділ", 14),
                                 ("Дія: статус", 18), ("Тип броні", 12)]:
            tk.Label(ff, text=f"{col_name}:", bg=CLR["bg"],
                     font=("Segoe UI", 8)).pack(side="left", padx=(4, 0))
            var = tk.StringVar(value="Всі")
            cb = ttk.Combobox(ff, textvariable=var, state="readonly",
                              width=width, font=("Segoe UI", 9))
            cb.pack(side="left", padx=(2, 4))
            cb.bind("<<ComboboxSelected>>", lambda *_, v=var: self._apply_filters())
            self._filter_vars[col_name] = (var, cb)

        tk.Button(ff, text="✕ Скинути", command=self._reset_filters,
                  bg=CLR["grey_fg"], fg="white", font=("Segoe UI", 8),
                  relief="flat", padx=6, pady=1, cursor="hand2").pack(side="left", padx=4)

        sf = tk.Frame(f, bg=CLR["bg"]); sf.pack(fill="x", padx=10, pady=(0, 6))
        tk.Label(sf, text="🔍 Пошук (Ctrl+F):", bg=CLR["bg"],
                 font=("Segoe UI", 9)).pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace("w", lambda *_: self._apply_filters())
        self.search_entry = tk.Entry(sf, textvariable=self.search_var,
                 font=("Segoe UI", 10), width=32)
        self.search_entry.pack(side="left", padx=6)
        tk.Button(sf, text="📥 Експорт CSV (Ctrl+E)",
                  command=lambda: self._export_csv(self.tree),
                  bg=CLR["dark_blue"], fg="white", font=("Segoe UI", 9),
                  relief="flat", padx=8, pady=2, cursor="hand2").pack(side="right")

        # Легенда кольорів внизу
        lg = tk.Frame(f, bg=CLR["bg"]); lg.pack(fill="x", padx=10, pady=(0, 6))
        tk.Label(lg, text="Підказка:", bg=CLR["bg"], fg=CLR["grey_fg"],
                 font=("Segoe UI", 8, "italic")).pack(side="left")
        for bg, fg, txt in [
            (CLR["green_bg"], CLR["green_fg"], "OK"),
            (CLR["red_bg"],   CLR["red_fg"],   "Донарахувати"),
            ("#FFF2CC",       "#806000",       "Бронь < 60 дн."),
            (CLR["red_bg"],   CLR["red_fg"],   "Прострочено"),
            ("#FFFFFF",       CLR["orange"],   "Лише в Дії"),
        ]:
            tk.Label(lg, text=f" {txt} ", bg=bg, fg=fg,
                     font=("Segoe UI", 8, "bold"),
                     relief="solid", borderwidth=1).pack(side="left", padx=3)
        tk.Label(lg, text="  Подвійний клік на рядку — деталі  │  "
                  "Правий клік — копіювати",
                  bg=CLR["bg"], fg=CLR["grey_fg"],
                  font=("Segoe UI", 8, "italic")).pack(side="right")

        # Контекстне меню та подвійний клік
        self._setup_tree_interactions(self.tree, on_double=self._on_main_row_double)

    def _build_summary_tab(self):
        f = self.tab_summary
        tk.Label(f, text="ПІДСУМКИ", bg=CLR["dark_blue"], fg="white",
                 font=("Segoe UI", 12, "bold")).pack(
                     fill="x", padx=10, pady=(10, 0), ipady=8)
        self.tree_summary = self._make_tree(
            f, ("Показник", "Значення"), [370, 240], height=14)
        self._setup_tree_interactions(self.tree_summary)
        bf = tk.Frame(f, bg=CLR["bg"]); bf.pack(fill="x", padx=10, pady=(0, 6))
        tk.Button(bf, text="📥 Експорт CSV",
                  command=lambda: self._export_csv(self.tree_summary),
                  bg=CLR["dark_blue"], fg="white", font=("Segoe UI", 9),
                  relief="flat", padx=8, pady=2, cursor="hand2").pack(side="right")

    def _build_avg_tab(self):
        f = self.tab_avg
        self._avg_title_var = tk.StringVar(value="АНАЛІТИКА")
        tk.Label(f, textvariable=self._avg_title_var,
                 bg=CLR["dark_blue"], fg="white",
                 font=("Segoe UI", 12, "bold")).pack(
                     fill="x", padx=10, pady=(10, 0), ipady=8)

        # Картки військового обліку (показуємо коли є дані Дії)
        self._dia_cards_frame = tk.Frame(f, bg=CLR["bg"])
        self.dia_card_total    = self._card(self._dia_cards_frame, "Всього у Дії",       "—", CLR["mid_blue"])
        self.dia_card_liable   = self._card(self._dia_cards_frame, "Військовозобовʼязані","—", CLR["dark_blue"])
        self.dia_card_bron     = self._card(self._dia_cards_frame, "Заброньовані",       "—", CLR["green_fg"])
        self.dia_card_temp     = self._card(self._dia_cards_frame, "Тимчас. забр.",      "—", "#2980B9")
        self.dia_card_defer    = self._card(self._dia_cards_frame, "Відстрочка",         "—", CLR["orange"])
        self.dia_card_pct      = self._card(self._dia_cards_frame, "% забронь.",         "—", CLR["purple"])
        self.dia_card_remain   = self._card(self._dia_cards_frame, "Залишок місць",      "—", CLR["red_fg"])

        self._dia_limit_info_var = tk.StringVar(value="")
        self._dia_limit_lbl = tk.Label(f, textvariable=self._dia_limit_info_var,
                 bg=CLR["bg"], fg=CLR["dark_blue"],
                 font=("Segoe UI", 9, "bold"), anchor="w")

        self.tree_avg = self._make_tree(
            f, ("Показник", "Значення", "Поріг / Деталі", "Відповідність"),
            [340, 170, 290, 170], height=22)
        self._setup_tree_interactions(self.tree_avg)
        bf = tk.Frame(f, bg=CLR["bg"]); bf.pack(fill="x", padx=10, pady=(0, 6))
        tk.Button(bf, text="📥 Експорт CSV",
                  command=lambda: self._export_csv(self.tree_avg),
                  bg=CLR["dark_blue"], fg="white", font=("Segoe UI", 9),
                  relief="flat", padx=8, pady=2, cursor="hand2").pack(side="right")

    # ── Завантаження файлу ────────────────────────────────

    def _open_file(self):
        path = filedialog.askopenfilename(
            title="Оберіть xlsx-файл",
            filetypes=[("Excel файли", "*.xlsx"), ("Всі файли", "*.*")])
        if path: self._load_file(path)

    def _load_file(self, path):
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            messagebox.showerror("Помилка відкриття", str(e)); return
        self.current_file = path
        self._wb = wb
        self._add_recent(path, "recent_files")
        candidates = [s for s in wb.sheetnames if s != OUTPUT_SHEET]
        self._sheet_cb.configure(values=candidates, state="readonly")
        self._pick_sheet(wb.sheetnames, self._analyse_sheet)

    def _on_sheet_change(self, _=None):
        sheet = self._sheet_var.get()
        if sheet and self._wb:
            self._analyse_sheet(sheet)

    # ── Ліміт бронювання ─────────────────────────────────

    def _on_bron_limit_change(self):
        try:
            pct = self._bron_limit_var.get()
        except tk.TclError:
            return
        self.settings["bron_limit_pct"] = pct
        save_settings(self.settings)
        if self.dia_data:
            self.dia_stats = analyse_dia(self.dia_data, bron_limit_pct=pct)
            self._populate_dia()
            if self.data:
                self._populate()
            else:
                for item in self.tree_avg.get_children(): self.tree_avg.delete(item)
                for label, val, minimum, status, tag in self._build_dia_analytics_rows():
                    self.tree_avg.insert("", "end",
                        values=(label, val, minimum, status), tags=(tag,))

    # ── Завантаження файлу Дії ────────────────────────────

    def _open_dia_file(self):
        path = filedialog.askopenfilename(
            title="Оберіть xlsx-файл з Дії (Реєстр військовозобовʼязаних)",
            filetypes=[("Excel файли", "*.xlsx"), ("Всі файли", "*.*")])
        if not path:
            return
        try:
            self.dia_data = read_dia_file(path)
        except Exception as e:
            messagebox.showerror("Помилка", f"Не вдалося прочитати файл Дії:\n{e}")
            return
        if not self.dia_data:
            messagebox.showwarning("Увага", "Файл не містить даних про працівників.")
            return
        pct = self._bron_limit_var.get()
        self.dia_stats = analyse_dia(self.dia_data, bron_limit_pct=pct)
        self._populate_dia()
        self._add_recent(path, "recent_dia")
        # Оновити фільтри і підсумки
        self._update_filter_options()
        if self.data:
            self._populate()
        else:
            # Тільки Дія, без ЗП — оновлюємо лише головну таблицю
            self._apply_filters()
            # І аналітику (тільки Дія)
            for item in self.tree_avg.get_children(): self.tree_avg.delete(item)
            for label, val, minimum, status, tag in self._build_dia_analytics_rows():
                self.tree_avg.insert("", "end",
                    values=(label, val, minimum, status), tags=(tag,))
        self.nb.select(self.tab_main)
        self.status_var.set(
            f"📂 Дія: {os.path.basename(path)}  │  "
            f"{self.dia_stats['total']} осіб  │  "
            f"{self.dia_stats['liable']} в/з  │  "
            f"{self.dia_stats['excluded']} виключених")

    def _populate_dia(self):
        ds = self.dia_stats
        # Показати картки та інфо про ліміт у вкладці "Аналітика"
        self._dia_cards_frame.pack(fill="x", padx=10, pady=(8, 6),
                                    before=self.tree_avg.master)
        self._dia_limit_lbl.pack(fill="x", padx=14, pady=(0, 4),
                                  before=self.tree_avg.master)

        self.dia_card_total.config(text=str(ds["total"]))
        self.dia_card_liable.config(text=str(ds["liable"]))
        self.dia_card_bron.config(text=str(ds["bronied"]))
        self.dia_card_temp.config(text=str(ds["temp_bron"]))
        self.dia_card_defer.config(text=str(ds["deferred"]))
        self.dia_card_pct.config(text=f"{ds['pct_bron_of_liable']:.1f}%")
        self.dia_card_remain.config(text=str(ds["remaining"]))

        self._dia_limit_info_var.set(
            f"Макс. дозволено: {ds['max_allowed']} з {ds['liable']} "
            f"({ds['bron_limit_pct']}%)  │  "
            f"Фактично: {ds['all_bron']}  │  "
            f"{'✓ В межах ліміту' if ds['over_limit'] == 0 else '✗ ПЕРЕВИЩЕНО!'}")

        # Попередження переліміту
        warnings = []
        if ds["over_limit"] > 0:
            warnings.append(
                f"⚠ ПЕРЕЛІМІТ! Заброньовано {ds['all_bron']}, "
                f"дозволено {ds['max_allowed']} ({ds['bron_limit_pct']}%). "
                f"Перевищення на {ds['over_limit']} осіб!")
        if ds["expired"]:
            names = ", ".join(p["pib"] for p, _ in ds["expired"])
            warnings.append(f"⚠ ПРОСТРОЧЕНА БРОНЬ ({len(ds['expired'])}): {names}")
        if ds["expiring_soon"]:
            parts = [f"{p['pib']} ({d} дн.)" for p, d in ds["expiring_soon"]]
            warnings.append(f"⏰ Закінчується скоро ({len(ds['expiring_soon'])}): {', '.join(parts)}")

        self._dia_warn_frame.pack_forget()
        if warnings:
            self._dia_warn_var.set("\n".join(warnings))
            self._dia_warn_frame.pack(fill="x", padx=10, pady=(0, 2),
                                       before=self.nb)
        else:
            self._dia_warn_var.set("")

    def _build_dia_analytics_rows(self):
        """Будує рядки аналітики Дії у форматі (label, val, detail, status, tag)
        для вставки в tree_avg."""
        ds = self.dia_stats
        if not ds:
            return []
        now_str = datetime.now().strftime("%d.%m.%Y")
        rows = [
            ("═══ ВІЙСЬКОВИЙ ОБЛІК (ДІЯ) ═══", "", "", "", ""),
            ("Всього працівників у файлі Дії", str(ds["total"]), "", "—", "zebra"),
            ("Виключені з військового обліку", str(ds["excluded"]),
             "Не враховуються", "—", "excl"),
            ("Військовозобовʼязані (облікові)", str(ds["liable"]),
             f"= {ds['total']} − {ds['excluded']}", "—", "zebra"),

            ("─── Бронювання ───", "", "", "", ""),
            ("Заброньовані (постійно)", str(ds["bronied"]),
             f"{ds['bronied']/ds['liable']*100:.1f}% від в/з" if ds["liable"] else "—",
             "—", "bron"),
            ("Заброньовані (тимчасово)", str(ds["temp_bron"]),
             f"{ds['temp_bron']/ds['liable']*100:.1f}% від в/з" if ds["liable"] else "—",
             "—", "defer" if ds["temp_bron"] > 0 else "zebra"),
            ("ВСЬОГО ЗАБРОНЬОВАНИХ", str(ds["all_bron"]),
             f"{ds['pct_bron_of_liable']:.1f}% від {ds['liable']} в/з",
             "✓ В МЕЖАХ" if ds["over_limit"] == 0 else "✗ ПЕРЕВИЩЕНО",
             "ok" if ds["over_limit"] == 0 else "need"),
            ("Мають відстрочку", str(ds["deferred"]),
             f"{ds['pct_defer']:.1f}% від в/з" if ds["liable"] else "—",
             "—", "defer"),
            ("Без статусу / не заброньовані", str(ds["not_bron"]),
             f"{ds['pct_not_bron']:.1f}% від в/з" if ds["liable"] else "—",
             "—", "zebra"),

            ("─── Ліміт бронювання ───", "", "", "", ""),
            ("Встановлений ліміт", f"{ds['bron_limit_pct']}%",
             f"Від {ds['liable']} в/з", "—", "zebra"),
            ("Максимально дозволено", str(ds["max_allowed"]),
             f"= {ds['liable']} × {ds['bron_limit_pct']}%", "—", ""),
            ("Фактично заброньовано", str(ds["all_bron"]),
             f"пост. {ds['bronied']} + тимч. {ds['temp_bron']}", "—", "zebra"),
        ]
        if ds["over_limit"] > 0:
            rows.append(("⚠ ПЕРЕВИЩЕННЯ ЛІМІТУ", f"+{ds['over_limit']} осіб",
                         f"Заброн. {ds['all_bron']}, ліміт {ds['max_allowed']}",
                         "✗ ПЕРЕВИЩЕНО", "need"))
        else:
            rows.append(("Залишок місць для бронювання", str(ds["remaining"]),
                         f"= {ds['max_allowed']} − {ds['all_bron']}",
                         "✓ ДОСТУПНО" if ds["remaining"] > 0 else "—",
                         "ok" if ds["remaining"] > 0 else "zebra"))

        # Строки
        rows.append(("─── Строки бронювання ───", "", f"Станом на {now_str}", "", ""))
        now = datetime.now()
        if ds["earliest_exp"]:
            days = (ds["earliest_exp"] - now).days
            rows.append(("Найближчий термін закінчення",
                         ds["earliest_exp"].strftime("%d.%m.%Y"),
                         f"через {days} дн.", "—",
                         "need" if days <= 30 else "zebra"))
        if ds["latest_exp"]:
            days = (ds["latest_exp"] - now).days
            rows.append(("Найпізніший термін закінчення",
                         ds["latest_exp"].strftime("%d.%m.%Y"),
                         f"через {days} дн.", "—", "zebra"))
        if ds["expired"]:
            names = ", ".join(p["pib"] for p, _ in ds["expired"])
            rows.append(("⚠ ПРОСТРОЧЕНА БРОНЬ", str(len(ds["expired"])),
                         names, "✗ ТЕРМІНОВО", "need"))
        if ds["expiring_soon"]:
            parts = ", ".join(f"{p['pib']} ({d}дн.)" for p, d in ds["expiring_soon"])
            rows.append(("⏰ Закінчується < 60 днів", str(len(ds["expiring_soon"])),
                         parts, "⚠ УВАГА", "defer"))

        # Перехрест із ЗП
        if self.data:
            rows.append(("═══ ПЕРЕХРЕСТ ЗП ↔ ДІЯ ═══", "", "", "", ""))
            zp_names = {_norm_name(r["name"]) for r in self.data}
            dia_bron_names = {p["pib_key"] for p in ds["all_bron_persons"]}
            zp_bron_names = {_norm_name(r["name"]) for r in self.data if is_bron_yes(r["bron"])}
            in_both       = dia_bron_names & zp_names
            in_dia_not_zp = dia_bron_names - zp_names
            in_zp_not_dia = zp_bron_names - dia_bron_names

            rows.append(("Працівники у ЗП-файлі", str(len(zp_names)),
                         "", "—", "zebra"))
            rows.append(("Заброньовані у ЗП-файлі", str(len(zp_bron_names)),
                         "", "—", ""))
            rows.append(("Збіг: бронь у Дії + у ЗП", str(len(in_both)),
                         "В обох джерелах", "✓" if in_both else "—",
                         "ok" if in_both else "zebra"))
            if in_dia_not_zp:
                names_list = ", ".join(p["pib"] for p in ds["all_bron_persons"]
                                        if p["pib_key"] in in_dia_not_zp)
                rows.append(("⚠ Бронь у Дії, НЕМАЄ у ЗП",
                             str(len(in_dia_not_zp)), names_list,
                             "✗ ПЕРЕВІРИТИ", "need"))
            else:
                rows.append(("Бронь у Дії, немає у ЗП", "0",
                             "Всі знайдені у ЗП", "✓", "ok"))
            if in_zp_not_dia:
                names_list = ", ".join(r["name"] for r in self.data
                                        if is_bron_yes(r["bron"])
                                        and _norm_name(r["name"]) in in_zp_not_dia)
                rows.append(("⚠ Бронь у ЗП, НЕМАЄ у Дії",
                             str(len(in_zp_not_dia)), names_list,
                             "✗ ПЕРЕВІРИТИ", "need"))
            else:
                rows.append(("Бронь у ЗП, немає у Дії", "0",
                             "Всі знайдені у Дії", "✓", "ok"))

            threshold = self.stats.get("threshold", 0)
            if threshold > 0:
                rows.append(("─── Перевірка ЗП порогу для броні з Дії ───", "",
                             f"Поріг: {threshold:,.2f} грн", "", ""))
                zp_by_name = {_norm_name(r["name"]): r for r in self.data}
                ok_n, fail_n, fails = 0, 0, []
                for p in ds["all_bron_persons"]:
                    zp_rec = zp_by_name.get(p["pib_key"])
                    if zp_rec:
                        if zp_rec["suma"] >= threshold:
                            ok_n += 1
                        else:
                            fail_n += 1
                            fails.append(f"{p['pib']} ({zp_rec['suma']:,.0f})")
                rows.append(("ЗП ≥ порогу", str(ok_n),
                             f"з {len(in_both)} знайдених",
                             "✓ ВІДПОВІДАЄ" if ok_n > 0 else "—",
                             "ok" if ok_n > 0 else "zebra"))
                if fail_n > 0:
                    rows.append(("⚠ ЗП < порогу", str(fail_n),
                                 "; ".join(fails),
                                 "✗ ДОНАРАХУВАТИ", "need"))
                else:
                    rows.append(("ЗП < порогу", "0",
                                 "Всі відповідають", "✓", "ok"))
        return rows

    def _hide_dia_widgets(self):
        """Сховати картки/інфо Дії якщо немає даних."""
        self._dia_cards_frame.pack_forget()
        self._dia_limit_lbl.pack_forget()
        self._dia_warn_frame.pack_forget()

    # ── Аналіз ────────────────────────────────────────────

    def _analyse_sheet(self, sheet_name):
        self.current_sheet = sheet_name
        self._sheet_var.set(sheet_name)
        rows, headers = read_sheet(self._wb[sheet_name])
        min_sal = self._get_min_sal()
        coeff   = self._get_coeff()
        self.data, self.stats = analyse(rows, headers, min_sal, coeff)
        s = self.stats
        self.status_var.set(
            f"📂  {os.path.basename(self.current_file)}  │  Аркуш: «{sheet_name}»  │  "
            f"{s['unique_n']} пр-ків  │  {s['bron_count']} заброньованих  │  "
            f"Поріг: {s['threshold']:,.2f} грн  │  "
            f"Працівник: «{s['cols']['employee'] or '?'}»  │  "
            f"Сума: «{s['cols']['amount'] or '?'}»"
        )
        self.btn_save.config(state="normal")
        self._avg_title_var.set(
            f"ЗП ЗАБРОНЬОВАНИХ vs ПОРІГ {s['threshold']:,.2f} грн"
            f"  (Мін. {min_sal:,} × {coeff})")
        self._update_filter_options()
        self._populate()

    def _reanalyse(self):
        if self.current_sheet and self._wb:
            self.settings["min_salary"] = self._get_min_sal()
            save_settings(self.settings)
            self._analyse_sheet(self.current_sheet)

    # ── Фільтри головної таблиці ─────────────────────────

    def _build_merged_rows(self):
        """Об'єднує ЗП-дані з даними Дії за нормалізованим ПІБ.
        Повертає список dict з усіма полями."""
        merged = []
        dia_by_name = {}
        if self.dia_data:
            for p in self.dia_data:
                if DIA_EXCLUDED in p.get("note", ""):
                    continue
                dia_by_name[p["pib_key"]] = p

        seen_keys = set()
        for r in self.data or []:
            key = _norm_name(r["name"])
            seen_keys.add(key)
            dia = dia_by_name.get(key)
            dia_status = dia["status"] if dia else ""
            dia_type   = dia["type"] if dia else ""
            dia_note   = dia.get("note", "") if dia else ""
            merged.append({
                **r,
                "in_zp": True,
                "in_dia": dia is not None,
                "dia_status": dia_status,
                "dia_type": "Тимчасово" if dia_type == "тимчасово"
                            else ("Постійно" if dia_type == "основний" and dia_status == DIA_BRON else ""),
                "dia_term": dia_note,
            })

        # Працівники з Дії, яких немає в ЗП-файлі
        for key, p in dia_by_name.items():
            if key in seen_keys:
                continue
            merged.append({
                "name":   p["pib"],
                "dept":   "",
                "bron":   "—",
                "suma":   0.0,
                "status": "Немає в ЗП-файлі",
                "donar":  "",
                "in_zp":  False,
                "in_dia": True,
                "dia_status": p["status"],
                "dia_type":   "Тимчасово" if p["type"] == "тимчасово"
                              else ("Постійно" if p["status"] == DIA_BRON else ""),
                "dia_term":   p.get("note", ""),
            })
        return merged

    def _update_filter_options(self):
        """Оновити варіанти в комбобоксах фільтрів."""
        if not self.data and not self.dia_data:
            return
        merged = self._build_merged_rows()
        bron_vals = sorted(set(str(r["bron"] or "—") for r in merged))
        stat_vals = sorted(set(r["status"] for r in merged))
        dept_vals = sorted(set(r["dept"] for r in merged if r["dept"]))
        dia_stat_vals = sorted(set(r["dia_status"] for r in merged if r["dia_status"]))
        dia_type_vals = sorted(set(r["dia_type"] for r in merged if r["dia_type"]))

        for col_name, values in [("Бронь",       bron_vals),
                                  ("Статус",      stat_vals),
                                  ("Підрозділ",   dept_vals),
                                  ("Дія: статус", dia_stat_vals),
                                  ("Тип броні",   dia_type_vals)]:
            var, cb = self._filter_vars[col_name]
            cb.configure(values=["Всі"] + values)
            var.set("Всі")

    def _reset_filters(self):
        for var, _ in self._filter_vars.values():
            var.set("Всі")
        self.search_var.set("")
        self._apply_filters()

    def _apply_filters(self):
        q = self.search_var.get().lower()
        f_bron     = self._filter_vars["Бронь"][0].get()
        f_stat     = self._filter_vars["Статус"][0].get()
        f_dept     = self._filter_vars["Підрозділ"][0].get()
        f_dia_st   = self._filter_vars["Дія: статус"][0].get()
        f_dia_type = self._filter_vars["Тип броні"][0].get()

        for item in self.tree.get_children():
            self.tree.delete(item)
        merged = self._build_merged_rows()
        num = 0
        for r in merged:
            if q and q not in r["name"].lower() and q not in r["dept"].lower():
                continue
            if f_bron != "Всі" and str(r["bron"] or "—") != f_bron:
                continue
            if f_stat != "Всі" and r["status"] != f_stat:
                continue
            if f_dept != "Всі" and r["dept"] != f_dept:
                continue
            if f_dia_st != "Всі" and r["dia_status"] != f_dia_st:
                continue
            if f_dia_type != "Всі" and r["dia_type"] != f_dia_type:
                continue
            num += 1
            donar_str = f"{r['donar']:,.2f}" if r["donar"] != "" else ""
            suma_str  = f"{r['suma']:,.2f}" if r["in_zp"] else "—"
            # Перевірка терміну броні
            term_tag = None
            if r["dia_status"] == DIA_BRON and r["dia_term"]:
                dt = _parse_dia_date(r["dia_term"])
                if dt:
                    days = (dt - datetime.now()).days
                    if days < 0:
                        term_tag = "expired"
                    elif days <= 60:
                        term_tag = "expiring"
            values = (num, r["name"], r["dept"], r["bron"],
                      suma_str, r["status"], donar_str,
                      r["dia_status"] or "—", r["dia_type"] or "—",
                      r["dia_term"] or "—")
            if r["status"] == STATUS_NEED:   tag = "need"
            elif term_tag == "expired":      tag = "expired"
            elif term_tag == "expiring":     tag = "expiring"
            elif r["status"] == STATUS_OK:   tag = "ok"
            elif not r["in_zp"]:             tag = "defer"
            elif num % 2 == 0:               tag = "zebra"
            else:                            tag = ""
            self.tree.insert("", "end", values=values, tags=(tag,))

    # ── Заповнення таблиць ────────────────────────────────

    def _populate(self):
        d, s      = self.data, self.stats
        threshold = s["threshold"]

        self.card_total.config(text=str(s["unique_n"]))
        self.card_bron.config( text=str(s["bron_count"]))
        self.card_ok.config(   text=str(s["ok_count"]))
        self.card_need.config( text=str(s["need_count"]))
        self.card_donar.config(text=f"{s['donar_total']:,.2f} грн")

        # Головна таблиця — через фільтри
        self._apply_filters()

        cols = s["cols"]
        for item in self.tree_summary.get_children(): self.tree_summary.delete(item)
        summary_rows = [
            ("Всього унікальних працівників",       str(s["unique_n"])),
            ("Всього заброньованих",                str(s["bron_count"])),
            ("Поріг виконано",                      str(s["ok_count"])),
            ("Потрібно донарахувати (осіб)",         str(s["need_count"])),
            ("Загальна сума до донарахування",      f"{s['donar_total']:,.2f} грн"),
            ("── Параметри розрахунку ──",           ""),
            ("Мінімальна ЗП (база)",                f"{s['min_salary']:,} грн"),
            (f"Коефіцієнт",                         f"×{s['coeff']}  ({self._coeff_var.get()})"),
            ("Поріг для заброньованих",             f"{threshold:,.2f} грн"),
            ("── Джерело ──",                        ""),
            ("Аркуш",                               self.current_sheet or "—"),
            ("Колонка «Бронь»",                     cols["bron"]     or "не знайдено"),
            ("Колонка «Сума»",                      cols["amount"]   or "не знайдено"),
            ("Колонка «Працівник»",                 cols["employee"] or "не знайдено"),
            ("Значення фільтра нарахувань",         s["filter_val"]  or "не знайдено"),
        ]
        # Якщо є дані Дії — додати аналітику
        if self.dia_stats:
            ds = self.dia_stats
            summary_rows.append(("── Військовий облік (Дія) ──", ""))
            summary_rows.append(("Всього у файлі Дії",         str(ds["total"])))
            summary_rows.append(("Військовозобовʼязані",       str(ds["liable"])))
            summary_rows.append(("Заброньовані (постійно)",    str(ds["bronied"])))
            summary_rows.append(("Заброньовані (тимчасово)",   str(ds["temp_bron"])))
            summary_rows.append(("Всього заброньованих",       str(ds["all_bron"])))
            summary_rows.append(("Мають відстрочку",           str(ds["deferred"])))
            summary_rows.append(("Виключені з обліку",         str(ds["excluded"])))
            summary_rows.append((f"Ліміт бронювання ({ds['bron_limit_pct']}%)",
                                 f"макс. {ds['max_allowed']} з {ds['liable']}"))
            summary_rows.append(("Залишок місць",              str(ds["remaining"])))
            summary_rows.append(("% заброньованих від в/з",    f"{ds['pct_bron_of_liable']:.1f}%"))

        for i, (k, v) in enumerate(summary_rows):
            tag = "red"   if "дон" in k.lower() and s["need_count"] > 0 else \
                  "zebra" if i % 2 == 0 else ""
            self.tree_summary.insert("", "end", values=(k, v), tags=(tag,))

        for item in self.tree_avg.get_children(): self.tree_avg.delete(item)
        avg_rows = [
            ("Загальна сума нарахувань (всі пр-ки)",
             f"{s['total_sum']:,.2f} грн", "—", "—", "zebra"),
            ("Кількість заброньованих",
             str(s["bron_count"]), "—", "—", ""),
            ("Сума нарахувань заброньованих",
             f"{s['bron_sum']:,.2f} грн", "—", "—", "zebra"),
            ("Необхідна загальна сума (поріг × N)",
             f"{s['need_total']:,.2f} грн", "—", "—", ""),
            ("Середня ЗП (всі працівники)",
             f"{s['avg_all']:,.2f} грн",
             f"{threshold:,.2f}",
             "✓ ВІДПОВІДАЄ" if s["avg_all"] >= threshold else "✗ НЕ ВІДПОВІДАЄ",
             "ok" if s["avg_all"] >= threshold else "need"),
            ("Дефіцит / Профіцит заброньованих",
             f"{s['deficit']:+,.2f} грн", "(факт − поріг×N)",
             "✓ ПРОФІЦИТ" if s["deficit"] >= 0 else "✗ ДЕФІЦИТ",
             "surplus" if s["deficit"] >= 0 else "deficit"),
        ]
        # Аналітика по даних Дії — додається в кінець
        if self.dia_stats:
            avg_rows.extend(self._build_dia_analytics_rows())

        for label, val, minimum, status, tag in avg_rows:
            self.tree_avg.insert("", "end",
                values=(label, val, minimum, status), tags=(tag,))

    # ── Експорт CSV ─────────────────────────────────────

    def _export_csv(self, tree):
        cols = tree["columns"]
        items = tree.get_children()
        if not items:
            messagebox.showinfo("", "Таблиця порожня — нема що експортувати."); return
        path = filedialog.asksaveasfilename(
            title="Зберегти CSV",
            defaultextension=".csv",
            filetypes=[("CSV файли", "*.csv"), ("Всі файли", "*.*")])
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerow(cols)
                for item in items:
                    writer.writerow(tree.item(item, "values"))
            messagebox.showinfo("Готово", f"CSV збережено:\n{path}")
        except Exception as e:
            messagebox.showerror("Помилка", str(e))

    # ── Збереження ────────────────────────────────────────

    def _save_excel(self):
        if not self._wb or not self.current_file: return
        try:
            # Очищаємо дати у вихідному аркуші (видаляємо " 00:00:00")
            cleaned = 0
            if self.current_sheet and self.current_sheet in self._wb.sheetnames:
                cleaned = clean_dates_in_sheet(self._wb[self.current_sheet])
            write_excel(self._wb, self.data, self.stats, self.current_sheet)
            self._wb.save(self.current_file)
            extra = f"\n(Дати очищено: {cleaned} клітинок)" if cleaned else ""
            messagebox.showinfo("Готово",
                f"✅  Аркуш «{OUTPUT_SHEET}» оновлено!{extra}\n\n{self.current_file}")
        except PermissionError:
            messagebox.showerror("Помилка",
                "Файл відкритий в Excel.\nЗакрийте його і спробуйте знову.")
        except Exception as e:
            messagebox.showerror("Помилка збереження", str(e))


# ══════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = App()
    app.mainloop()
